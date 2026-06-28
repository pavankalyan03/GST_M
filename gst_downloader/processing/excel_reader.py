import os
import sys
import logging
import openpyxl
from gst_downloader import config

def read_irns_from_excel(filepath: str, logger: logging.Logger) -> list[dict]:
    """
    Read IRNs and invoice numbers from the Excel file.

    Returns:
        List of dicts, each with keys: irn, invoice_number, row
    """
    logger.info(f"Reading Excel file: {filepath}")

    if not os.path.exists(filepath):
        logger.error(f"Excel file not found: {filepath}")
        sys.exit(1)

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    logger.info(f"Sheet: '{ws.title}' (Parsing rows...)")

    records = []
    # Use iter_rows for blazing fast streaming of large files in read_only mode
    for row_idx, row in enumerate(
        ws.iter_rows(min_row=config.HEADER_ROW + 1, values_only=True),
        start=config.HEADER_ROW + 1
    ):
        # OpenPyXL row is a tuple of values (0-indexed). Columns are 1-indexed.
        if len(row) >= max(config.IRN_COLUMN, config.INVOICE_NUM_COLUMN, config.INVOICE_DATE_COLUMN):
            irn_val = row[config.IRN_COLUMN - 1]
            inv_val = row[config.INVOICE_NUM_COLUMN - 1]
            date_val = row[config.INVOICE_DATE_COLUMN - 1]
        else:
            # Row doesn't have enough columns, try to extract what we can
            irn_val = row[config.IRN_COLUMN - 1] if len(row) >= config.IRN_COLUMN else None
            inv_val = row[config.INVOICE_NUM_COLUMN - 1] if len(row) >= config.INVOICE_NUM_COLUMN else None
            date_val = row[config.INVOICE_DATE_COLUMN - 1] if len(row) >= config.INVOICE_DATE_COLUMN else None

        irn = str(irn_val).strip() if irn_val else None
        invoice_number = str(inv_val).strip() if inv_val else f"invoice_row_{row_idx}"
        
        # Openpyxl might parse dates as datetime objects. Formatting them to YYYY-MM-DD
        if hasattr(date_val, "strftime"):
            invoice_date = date_val.strftime("%Y-%m-%d")
        else:
            invoice_date = str(date_val).strip() if date_val else ""
            # Sanitize the date string to be safe for filenames
            invoice_date = invoice_date.replace("/", "-").replace("\\", "-").replace(" ", "_")

        if irn and irn.lower() not in ("none", ""):
            records.append({
                "irn": irn,
                "invoice_number": invoice_number,
                "invoice_date": invoice_date,
                "row": row_idx,
            })
        else:
            logger.warning(f"Row {row_idx}: Skipped - empty or missing IRN")

    wb.close()
    logger.info(f"Loaded {len(records)} valid IRN records")
    return records
