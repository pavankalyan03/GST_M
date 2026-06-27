import os
import openpyxl
from pathlib import Path
import logging

def preprocess_excel(input_path: str, output_path: str, logger: logging.Logger = None):
    if logger:
        logger.info(f"Preprocessing raw Excel file: {input_path}")
    else:
        print(f"Preprocessing raw Excel file: {input_path}")
        
    if not os.path.exists(input_path):
        msg = f"Raw Excel file not found: {input_path}"
        if logger:
            logger.error(msg)
        raise FileNotFoundError(msg)
        
    wb_in = openpyxl.load_workbook(input_path, data_only=True)
    if "B2B" not in wb_in.sheetnames:
        raise ValueError("The input Excel does not contain a 'B2B' sheet.")
        
    ws_in = wb_in["B2B"]
    
    rows = list(ws_in.iter_rows(min_row=5, max_row=6, values_only=True))
    if len(rows) < 2:
        raise ValueError("The 'B2B' sheet does not have enough rows for headers.")
        
    row5, row6 = rows[0], rows[1]
    
    # Construct combined header
    header = []
    for c5, c6 in zip(row5, row6):
        if c6 is not None and str(c6).strip():
            header.append(str(c6).strip())
        elif c5 is not None and str(c5).strip():
            header.append(str(c5).strip())
        else:
            header.append(f"Column_{len(header)}")
            
    # Find necessary column indices (0-indexed)
    try:
        rate_idx = header.index("Rate (%)")
        taxable_idx = header.index("Taxable Value (₹)")
    except ValueError as e:
        raise ValueError(f"Required column not found in headers: {e}")
        
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "B2B_Filtered"
    
    ws_out.append(header)
    
    valid_count = 0
    # Data starts from row 7
    for row in ws_in.iter_rows(min_row=7, values_only=True):
        if not any(row):  # skip completely empty rows
            continue
            
        rate_val = row[rate_idx]
        taxable_val = row[taxable_idx]
        
        # Safely parse numeric values
        try:
            rate = float(rate_val) if rate_val not in (None, "-", "") else 0.0
            taxable = float(taxable_val) if taxable_val not in (None, "-", "") else 0.0
        except ValueError:
            continue
            
        if rate >= 18.0 and taxable >= 500.0:
            ws_out.append(row)
            valid_count += 1
            
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb_out.save(output_path)
    wb_in.close()
    
    if logger:
        logger.info(f"Preprocessing complete. Extracted {valid_count} valid records into {output_path}")
    else:
        print(f"Preprocessing complete. Extracted {valid_count} valid records into {output_path}")
