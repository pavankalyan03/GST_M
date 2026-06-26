import os
import sys
import logging
import openpyxl
from gst_downloader import config

def read_irns_from_excel(filepath: str, logger: logging.Logger) -> list[dict]:
    """
    Read IRNs and invoice numbers from the Excel file.

    Returns:
        List of dicts, each with keys: irn, invoice_number, row
    """
    logger.info(f"Reading Excel file: {filepath}")

    if not os.path.exists(filepath):
        logger.error(f"Excel file not found: {filepath}")
        sys.exit(1)

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    logger.info(f"Sheet: '{ws.title}' | Rows: {ws.max_row} | Cols: {ws.max_column}")

    records = []
    for row_num in range(config.HEADER_ROW + 1, ws.max_row + 1):
        irn_val = ws.cell(row=row_num, column=config.IRN_COLUMN).value
        inv_val = ws.cell(row=row_num, column=config.INVOICE_NUM_COLUMN).value

        irn = str(irn_val).strip() if irn_val else None
        invoice_number = str(inv_val).strip() if inv_val else f"invoice_row_{row_num}"

        if irn and irn.lower() not in ("none", ""):
            records.append({
                "irn": irn,
                "invoice_number": invoice_number,
                "row": row_num,
            })
        else:
            logger.warning(f"Row {row_num}: Skipped - empty or missing IRN")

    wb.close()
    logger.info(f"Loaded {len(records)} valid IRN records")
    return records
